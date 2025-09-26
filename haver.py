import requests
import json
from haversine import haversine, Unit
from vehicle_list import vehicles
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import math 



API_URL = "https://"
THRESHOLD_M = 20000  # 1000 m üstündeki atlamalar yok sayılacak
TOKEN = "Token"  # JWT token ekle
 
def fetch_points(vehicle_value, start_date, end_date):
    headers = {
        "accept": "application/json",
        "authorization": TOKEN,
        "content-type": "application/json"
    }
 
    payload = {
        "start": start_date,
        "end": end_date,
        "vehicle": {
            "value": vehicle_value,
        },
        # "fleet_id": 4154 
        # "fleet_id": 4699  
        "fleet_id": 4017   
    }
 
    response = requests.post(API_URL, headers=headers, data=json.dumps(payload))
    # print(response.status_code)
    response.raise_for_status()
    data = response.json()

    # print("API Response:")    

    if "positions" not in data:
        raise ValueError("API response does not contain 'positions' key.")
    

    # Alan adlarını API'ye göre ayarla
    return [(float(item[0]), float(item[1])) for item in data["positions"]]
 
 
def calculate_total_m(points):
    total_m = 0.0
    last_point = None
 
    for index, point in enumerate(points):
        if last_point:
            dist = haversine(last_point, point, unit=Unit.METERS)
            if dist <= THRESHOLD_M:  # mantıksız atlamaları filtrele
                total_m += dist
                last_point = point
        if index == 0:
            last_point = point
 
    return total_m
 
 
def calculate(vehicle_value, start_date, end_date):
    points = fetch_points(vehicle_value, start_date, end_date)
    # print("points:", len(points))
    meters = calculate_total_m(points)
    # print(f"Toplam katedilen mesafe: {meters:.2f} metre")
    # print(f"Toplam katedilen mesafe: {meters/1000:.2f} km")
    return f"{meters/1000: .2f}"



index_sayısı = math.ceil(800 / 250)

for i in range(1, index_sayısı):
    print(f"\n▶️ mesaidışı-22-24-{i}.xlsx dosyası işleniyor...")

    df = pd.read_excel(f"mesaidışı-22-24-{i}.xlsx")
    data = []

    # openpyxl workbook
    wb = load_workbook(f"mesaidışı-22-24-{i}.xlsx")
    ws = wb.active

    # Excel başlıklarını kontrol et (gerekirse ekle)
    ws.cell(row=1, column=2, value="Plaka")
    ws.cell(row=1, column=8, value="KM(Mesai içi)")
    ws.cell(row=1, column=5, value="KM(Mesai dışı)")

    count = 0
    TOTAL = len(df)

    for index, row in df.iterrows():
        count = index+1

        if count <= (250*i):  #250*1 kadarını atlayacak şimdi burdaa
            continue 

        if count > (250*i+250):   #burda da 1000e geldiğinde break
            break 

        # --- Basit Progress Bar ---
        percent = ((count-(250*i)) / 250) * 100  # count=752 mesela eksi 750 = 2% gösterecek
        bar = "#" * int(percent // 2)
        print(f"\rİlerleme: [{bar:<50}] {count-(250*i)}/250 ({percent:.1f}%)", end="")



        plate = row["Plaka"]
        date = row["Tarih"]

        # Plakaya göre vehicle bul
        matches = next((v for v in vehicles if v["label"].startswith(plate)), None)
        if not matches:
            print(f"{plate} için eşleşen araç bulunamadı.")
            continue

        vehicle_value = matches["value"]
        print(f"{plate} ({vehicle_value}) için hesaplama...")

        # Tarihlere göre saat aralıkları
        start_in_date = f"{date} 05:00:00"
        end_in_date = f"{date} 21:00:00"
        start_out1_date = f"{date} 21:00:00"
        end_out1_date = f"{date} 23:59:00"
        start_out2_date = f"{date} 00:00:00"
        end_out2_date = f"{date} 05:00:00"

        try:
            distance_in_km = calculate(vehicle_value, start_in_date, end_in_date)
            distance_out1_km = calculate(vehicle_value, start_out1_date, end_out1_date)
            distance_out2_km = calculate(vehicle_value, start_out2_date, end_out2_date)

            distance_out_km = float(distance_out1_km) + float(distance_out2_km)

        except Exception as e:
            print(f"{plate} - {date} için hata oluştu: {e}")
            distance_in_km = None
            distance_out_km = None

        data.append({
            "Plaka": plate,
            "Tarih": date,
            "KM(Mesai içi)": distance_in_km,
            "KM(Mesai dışı)": distance_out_km
        })

        print(f"mesai dışı km {distance_out_km}")
        print(f"mesai içi km {distance_in_km}")

        ws.cell(row=index + 2, column=8, value=distance_in_km)
        ws.cell(row=index + 2, column=5, value=distance_out_km)
        ws.cell(row=index + 2, column=15, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


    wb.save(f"mesaidışı-22-24-{i+1}.xlsx")    # i=3ten başladı ve revize4 adında yeni dosya save edecek
    print("Revize dosya kaydedildi.")
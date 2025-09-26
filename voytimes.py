import time
from datetime import datetime, time
from vehicle_list import vehicles
import pandas as pd
from openpyxl import load_workbook
import requests
import json


API_URL = "https://"
ATS2_TOKEN = ""  # JWT token ekle

def getVoyageTime(vehicle_id, start_date, end_date):   
    headers = {
           "accept": "application/json",
           "Token": ATS2_TOKEN,
           "content-type": "application/json"
    }


    payload = {
        "data": {
          "params": {
            "report": 0,
            "server_side": "false",
            "start_date": start_date,
            "end_date": end_date,
            "vehicles": [
              {
                "value": vehicle_id,
              }
            ]
          }
        },
        "fleet_id": 4017
    }

    response = requests.post(API_URL, headers=headers, data=json.dumps(payload))
    response.raise_for_status()
    data = response.json()
    # print(data)
    
    return [ (item['start_date'], item['end_date']) for item in data["data"]]


def calculate(vehicle_id, date): 
  gece_bas = gece_bit = aksam_bas = aksam_bit = None
  # date = "22.09.2025"
  start_in_date = f"{date} 05:00:00"
  end_in_date = f"{date} 21:00:00"

  start_out1_date = f"{date} 21:00:00"
  end_out1_date = f"{date} 23:59:00"

  start_out2_date = f"{date} 00:00:00"
  end_out2_date = f"{date} 05:00:00"

  all_date_start = f"{date} 00:00:00"
  all_date_end = f"{date} 23:59:59"



  allVoyage = getVoyageTime(vehicle_id, all_date_start, all_date_end)
  day = getVoyageTime(vehicle_id, start_in_date, end_in_date)
  night1 = getVoyageTime(vehicle_id, start_out2_date, end_out2_date)
  night2 = getVoyageTime(vehicle_id, start_out1_date, end_out1_date)
  # print("ilk kontak", day[0][0])
  # print("son kontak kapanış", day[-1][1])

  start_times = [start for start, end in day]
  end_times = [end for start, end in day]
  daytimeVoyCount = len(start_times)
  allVoyageCount = len(allVoyage)
  # print(start_times)
  # print(end_times)
  # print(allVoyage[-1][1])


  if(allVoyageCount != daytimeVoyCount):
    night1_start_times = [start for start, end in night1]    #gece sefer varsa başlangıç 00:02
    night1_end_times = [end for start, end in night1]      ##bu ife bakar
    night2_start_times = [start for start, end in night2]    #gece mesaisi yoksa akşam varsa vardır    23.10    yine ife bakar
    night2_end_times = [end for start, end in night2]        #akşam varsa bitiş ve belki son bitiş
    if night1_start_times:
      gece_bas = night1_start_times[0]
      print("Gece yolculuğu başl:", night1_start_times[0])
      if not night2_start_times:
        gece_bit = night1_end_times[-1]
        print("Gece yolculuğu bitiş:", night1_end_times[-1])

    if night2_end_times:
      aksam_bit = night2_end_times[-1]
      print("Akşam yolculuğu bitş:", night2_end_times[-1])
      if not night1_start_times:
        aksam_bas = night2_start_times[0]
        print("Akşam yolculuğu baş:", night2_start_times[0])


  day_start = allVoyage[0][0]
  day_end = allVoyage[-1][1]
  print("tüm gün ilk kontak", day_start)
  print("tüm gün son kontak", day_end)

  return gece_bas, gece_bit, aksam_bas, aksam_bit, day_start, day_end


df = pd.read_excel(f"alagöz-mesaidışı-22-24.xlsx")
data = []
# openpyxl workbook
wb = load_workbook(f"alagöz-mesaidışı-22-24.xlsx")
ws = wb.active
# Excel başlıklarını kontrol et (gerekirse ekle)
ws.cell(row=1, column=2, value="Plaka")
ws.cell(row=1, column=4, value="Tarih")
ws.cell(row=1, column=12, value="Mesai Dışı İlk Kontak Açılışı")
ws.cell(row=1, column=13, value="Mesai Dışı Son Kontak Kapanışı")
ws.cell(row=1, column=14, value="Günün İlk Kontak Açılışı")
ws.cell(row=1, column=15, value="Günün Son Kontak Kapanışı")


count = 0

for index, row in df.iterrows():
    count = index+1
    # if count > (25):   #burda da 1000e geldiğinde break
    #         break 


    plate = row["Plaka"]
    date = row["Tarih"]

  # Plakaya göre vehicle bul
    matches = next((v for v in vehicles if v["label"].startswith(plate)), None)
    if not matches:
        print(f"{plate} için eşleşen araç bulunamadı.")
        continue  
    vehicle_value = matches["value"]
    print(f"{plate} ({vehicle_value}) için hesaplama...") 


    gece_bas = None
    gece_bit = None
    aksam_bas = None
    aksam_bit = None
    day_start = None
    day_end = None


    try:
        # gece_bas = calculate(vehicle_value, date,)
        # gece_bit = calculate(vehicle_value, date,)
        # aksam_bas = calculate(vehicle_value, date,)
        # aksam_bit = calculate(vehicle_value, date,)
        # day_start = calculate(vehicle_value, date,)
        # day_end = calculate(vehicle_value, date,)
        gece_bas, gece_bit, aksam_bas, aksam_bit, day_start, day_end = calculate(vehicle_value, date)


    except Exception as e:
          print(f"{plate} - {date} için hata oluştu: {e}")
          gece_bas = gece_bit = aksam_bas = aksam_bit = day_start = day_end = None

    data.append({
          "Plaka": plate,
          "Tarih": date,
          "Mesai Dışı İlk Kontak Açılışı": gece_bas if gece_bas else aksam_bas,
          "Mesai Dışı Son Kontak Kapanışı": aksam_bit if aksam_bit else gece_bit,
          "Günün İlk Kontak Açılışı": day_start,
          "Günün Son Kontak Kapanışı": day_end
      })
    
    print({
    "Plaka": plate,
    "Tarih": date,
    "Mesai Dışı İlk Kontak Açılışı": gece_bas if gece_bas else aksam_bas,
    "Mesai Dışı Son Kontak Kapanışı": aksam_bit if aksam_bit else gece_bit,
    "Günün İlk Kontak Açılışı": day_start,
    "Günün Son Kontak Kapanışı": day_end
})

    
    ws.cell(row=index + 2, column=2, value=plate)
    ws.cell(row=index + 2, column=4, value=date)
    ws.cell(row=index + 2, column=11, value=gece_bas if gece_bas else aksam_bas)
    ws.cell(row=index + 2, column=12, value=aksam_bit if aksam_bit else gece_bit)
    ws.cell(row=index + 2, column=13, value=day_start)
    ws.cell(row=index + 2, column=14, value=day_end)
  

wb.save(f"revizeSaatttt.xlsx")    # i=3ten başladı ve revize4 adında yeni dosya save edecek
print("Revize dosya kaydedildi.")